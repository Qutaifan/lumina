#!/usr/bin/env python3
"""Import Lumina listings from Project X Excel + photo folders.

- Reads `Grok Updated new.xlsx` (Listings sheet) as source of truth
- Matches photos from `Up to date listings/` by reference / photo folder
- Deduplicates by Ref No (keeps first occurrence)
- Copies/normalizes images into website/assets/listings/{ref}/
- Writes website/data/lumina-demo-leads.json (site feed) + import report
- NEVER publishes owner/guard phone numbers from the sheet

Usage:
  python scripts/import_project_x_listings.py
  python scripts/import_project_x_listings.py --max-photos 12
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageOps, UnidentifiedImageError

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(r"C:\Users\Yazan\Desktop\Last updated Project X")
SITE_DATA = ROOT / "website" / "data" / "lumina-demo-leads.json"
SITE_LISTINGS_DIR = ROOT / "website" / "assets" / "listings"
REPORT_PATH = ROOT / "website" / "data" / "import-report.json"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".heic", ".heif"}
SKIP_DIR_NAMES = {"water mark", "watermark", "raw", "__macosx"}
SKIP_NAME_PARTS = ("watermark", "water mark")


def norm_ref(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"none", "nan", "ref no (click for photos)"}:
        return None
    # strip leading zeros? keep as zero-padded 3 if pure int
    if re.fullmatch(r"\d+", s):
        return s.zfill(3) if len(s) <= 3 else s
    # refs like 090 already
    m = re.search(r"(\d{2,4})", s)
    return m.group(1).zfill(3) if m and len(m.group(1)) <= 3 else (m.group(1) if m else s)


def parse_int(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(round(float(value)))
    m = re.search(r"-?\d[\d,]*", str(value).replace(" ", ""))
    return int(m.group(0).replace(",", "")) if m else None


def parse_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    m = re.search(r"-?\d[\d,.]*", str(value).replace(" ", ""))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def slugify(*parts: str) -> str:
    raw = "-".join(p for p in parts if p)
    s = re.sub(r"[^a-zA-Z0-9]+", "-", raw.strip().lower()).strip("-")
    return s or "listing"


def folder_refs(folder_name: str) -> list[str]:
    """Expand folder names like 025-026-027 or 090-91-92-93-94-95-96 into ref list."""
    name = folder_name.strip()
    if re.fullmatch(r"\d+", name):
        return [norm_ref(name)]
    # split on - and _
    tokens = re.split(r"[-_]+", name)
    refs = []
    for t in tokens:
        t = t.strip()
        if not t:
            continue
        if re.fullmatch(r"\d+", t):
            refs.append(norm_ref(t))
    return refs


def is_image(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in IMAGE_EXTS


def should_skip_path(path: Path, root: Path) -> bool:
    rel_parts = [p.lower() for p in path.relative_to(root).parts]
    if any(p in SKIP_DIR_NAMES for p in rel_parts[:-1]):
        return True
    name = path.name.lower()
    if any(part in name for part in SKIP_NAME_PARTS):
        return True
    return False


def collect_folder_images(folder: Path) -> list[Path]:
    if not folder.is_dir():
        return []
    files = []
    for p in folder.rglob("*"):
        if not is_image(p):
            continue
        if should_skip_path(p, folder):
            continue
        files.append(p)
    # stable sort by path depth then name
    files.sort(key=lambda p: (len(p.parts), str(p).lower()))
    return files


def file_fingerprint(path: Path, chunk: int = 256 * 1024) -> str:
    h = hashlib.sha1()
    h.update(str(path.stat().st_size).encode())
    with path.open("rb") as f:
        h.update(f.read(chunk))
    return h.hexdigest()


def convert_to_jpeg(src: Path, dest: Path, max_side: int = 1920, quality: int = 82) -> bool:
    """Open image (incl. HEIC if codec available), orient, resize, save as JPEG."""
    try:
        with Image.open(src) as im:
            im = ImageOps.exif_transpose(im)
            if im.mode not in ("RGB", "L"):
                im = im.convert("RGB")
            elif im.mode == "L":
                im = im.convert("RGB")
            w, h = im.size
            scale = min(1.0, max_side / max(w, h))
            if scale < 1.0:
                im = im.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
            dest.parent.mkdir(parents=True, exist_ok=True)
            im.save(dest, "JPEG", quality=quality, optimize=True, progressive=True)
            return True
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        # HEIC without codec, corrupt, etc.
        print(f"  skip image {src.name}: {exc}")
        return False


def map_photo_folders(listings_root: Path) -> dict[str, Path]:
    """ref -> best matching folder path."""
    ref_to_folder: dict[str, Path] = {}
    if not listings_root.is_dir():
        return ref_to_folder

    # Prefer exact single-ref folders over multi-ref folders
    multi: list[tuple[Path, list[str]]] = []
    for d in sorted(listings_root.iterdir()):
        if not d.is_dir():
            continue
        refs = folder_refs(d.name)
        if not refs:
            continue
        if len(refs) == 1:
            ref_to_folder[refs[0]] = d
        else:
            multi.append((d, refs))

    for d, refs in multi:
        for r in refs:
            if r not in ref_to_folder:
                ref_to_folder[r] = d
    return ref_to_folder


def read_excel_listings(xlsx: Path) -> list[dict]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if "Listings" not in wb.sheetnames:
        raise SystemExit(f"No 'Listings' sheet in {xlsx}")
    ws = wb["Listings"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if h.lower() == n.lower():
                    return i
        for n in names:
            for i, h in enumerate(header):
                if n.lower() in h.lower():
                    return i
        return None

    idx = {
        "ref": col("Ref No (click for photos)", "Ref No", "Ref"),
        "photos": col("Photos"),
        "type": col("Property Type"),
        "tx": col("Transaction Type"),
        "location": col("Location"),
        "price": col("Price (JOD)", "Price"),
        "area": col("Area (SQM)", "Area"),
        "pps": col("Price per SQM (JOD)", "Price per SQM"),
        "beds": col("Bedrooms"),
        "baths": col("Bathrooms"),
        "floor": col("Floor"),
        "furnished": col("Furnished"),
        "status": col("Status"),
        "kitchen": col("Kitchen"),
        "living": col("Living Room"),
        "cooling": col("Cooling / Heating", "Cooling"),
        "outdoor": col("Outdoor"),
        "address": col("St. / Building No."),
        "notes": col("Notes"),
        "folder": col("Photo Folder"),
    }

    out = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        ref = norm_ref(row[idx["ref"]] if idx["ref"] is not None else None)
        if not ref:
            continue

        def g(key):
            i = idx[key]
            if i is None or i >= len(row):
                return None
            v = row[i]
            if isinstance(v, str):
                v = v.strip()
            return v

        folder = g("folder")
        folder = str(folder).strip() if folder not in (None, "") else ref

        out.append(
            {
                "ref": ref,
                "property_type": str(g("type") or "Residential").strip(),
                "transaction": str(g("tx") or "").strip(),
                "location": str(g("location") or "Amman").strip(),
                "price": parse_int(g("price")),
                "area": parse_float(g("area")),
                "price_per_sqm": parse_float(g("pps")),
                "bedrooms": parse_int(g("beds")),
                "bathrooms": parse_int(g("baths")),
                "floor": str(g("floor") or "").strip(),
                "furnished": str(g("furnished") or "").strip(),
                "status": str(g("status") or "").strip(),
                "kitchen": str(g("kitchen") or "").strip(),
                "living_room": str(g("living") or "").strip(),
                "cooling": str(g("cooling") or "").strip(),
                "outdoor": str(g("outdoor") or "").strip(),
                "address": str(g("address") or "").strip(),
                "notes": str(g("notes") or "").strip(),
                "photo_folder": folder,
                "photos_label": str(g("photos") or "").strip(),
            }
        )
    return out


def dedupe_by_ref(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    seen = {}
    dups = []
    ordered = []
    for r in rows:
        ref = r["ref"]
        if ref in seen:
            dups.append({"ref": ref, "kept": seen[ref], "dropped": r})
            continue
        seen[ref] = {
            "title": f"{r.get('location')} {r.get('property_type')}",
            "price": r.get("price"),
            "location": r.get("location"),
        }
        ordered.append(r)
    return ordered, dups


def build_title(row: dict) -> str:
    loc = row.get("location") or "Amman"
    ptype = row.get("property_type") or "Residence"
    tx = (row.get("transaction") or "").lower()
    if "sale" in tx and "rent" in tx:
        suffix = ""
    elif "sale" in tx:
        suffix = " for Sale"
    elif "rent" in tx:
        suffix = " for Rent"
    else:
        suffix = ""
    return f"{loc} {ptype}{suffix}".strip()


def description(row: dict) -> str:
    bits = []
    loc = row.get("location") or "Amman"
    ptype = (row.get("property_type") or "residence").lower()
    bits.append(f"A {ptype} in {loc}.")
    if row.get("area"):
        bits.append(f"Approx. {int(row['area'])} sqm.")
    if row.get("bedrooms"):
        bits.append(f"{row['bedrooms']} bedrooms.")
    if row.get("bathrooms"):
        bits.append(f"{row['bathrooms']} bathrooms.")
    if row.get("furnished"):
        bits.append(f"{row['furnished']}.")
    if row.get("floor"):
        bits.append(f"Floor: {row['floor']}.")
    if row.get("outdoor"):
        bits.append(str(row["outdoor"]).strip() + ".")
    if row.get("notes"):
        bits.append(str(row["notes"]).strip())
    bits.append("Details shared upon request. Availability and final particulars require verification.")
    return " ".join(bits)


def to_site_listing(row: dict, image_urls: list[str]) -> dict:
    price = row.get("price")
    ref = row["ref"]
    title = build_title(row)
    primary = image_urls[0] if image_urls else "/assets/images/hero-luxury-villa.jpg"
    area = row.get("area")
    size = int(round(area)) if isinstance(area, (int, float)) else None

    return {
        "id": f"lumina-{ref}",
        "ref": ref,
        "slug": slugify(ref, row.get("location") or "", row.get("property_type") or ""),
        "title": title,
        "description": description(row),
        "property_type": row.get("property_type") or "Residential",
        "transaction": row.get("transaction") or "",
        "location_area": f"{row.get('location') or 'Amman'} — Amman",
        "location": row.get("location") or "Amman",
        "price_jod_raw": price,
        "price_jod_test_margin": price,
        "size_sqm": size,
        "land_area_sqm": None,
        "bedrooms": str(row["bedrooms"]) if row.get("bedrooms") is not None else None,
        "bathrooms": str(row["bathrooms"]) if row.get("bathrooms") is not None else None,
        "floor": row.get("floor") or "",
        "furnished": row.get("furnished") or "",
        "status": row.get("status") or "",
        "kitchen": row.get("kitchen") or "",
        "living_room": row.get("living_room") or "",
        "cooling": row.get("cooling") or "",
        "outdoor": row.get("outdoor") or "",
        "address": row.get("address") or "",
        "notes": row.get("notes") or "",
        "source": "Lumina Portfolio",
        "source_url": f"property-details.html?id=lumina-{ref}",
        "contact_signal": "Enquire via Lumina WhatsApp",
        "verification_status": "PORTFOLIO / VERIFICATION REQUIRED",
        "image_url": primary,
        "images": image_urls,
        "photo_count": len(image_urls),
        "quality_score": 4 if image_urls else 2,
    }


def import_photos_for_ref(
    ref: str,
    folder: Path | None,
    dest_dir: Path,
    max_photos: int,
    seen_hashes: set[str],
) -> tuple[list[str], dict]:
    meta = {
        "ref": ref,
        "folder": str(folder) if folder else None,
        "source_images": 0,
        "imported": 0,
        "duplicates_skipped": 0,
        "failed": 0,
    }
    if not folder or not folder.is_dir():
        return [], meta

    sources = collect_folder_images(folder)
    meta["source_images"] = len(sources)

    dest_dir.mkdir(parents=True, exist_ok=True)
    # clear previous imports for this ref so re-runs are clean
    for old in dest_dir.glob("*"):
        if old.is_file():
            old.unlink()

    urls: list[str] = []
    for src in sources:
        if len(urls) >= max_photos:
            break
        fp = file_fingerprint(src)
        if fp in seen_hashes:
            meta["duplicates_skipped"] += 1
            continue
        idx = len(urls) + 1
        dest = dest_dir / f"{idx:02d}.jpg"
        if convert_to_jpeg(src, dest):
            seen_hashes.add(fp)
            urls.append(f"/assets/listings/{ref}/{dest.name}")
            meta["imported"] += 1
        else:
            meta["failed"] += 1
    return urls, meta


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    ap.add_argument("--max-photos", type=int, default=12)
    ap.add_argument("--limit", type=int, default=0, help="Optional cap for testing")
    args = ap.parse_args()

    source = args.source
    xlsx = source / "Grok Updated new.xlsx"
    photos_root = source / "Up to date listings"
    if not xlsx.exists():
        raise SystemExit(f"Excel not found: {xlsx}")
    if not photos_root.exists():
        raise SystemExit(f"Photo root not found: {photos_root}")

    print(f"Reading Excel: {xlsx}")
    raw = read_excel_listings(xlsx)
    print(f"Excel rows with refs: {len(raw)}")

    unique, dups = dedupe_by_ref(raw)
    print(f"Unique refs: {len(unique)} | Duplicate rows dropped: {len(dups)}")
    if dups:
        for d in dups[:15]:
            print(f"  DUP ref {d['ref']}: kept {d['kept']} dropped price={d['dropped'].get('price')}")

    if args.limit:
        unique = unique[: args.limit]

    ref_folders = map_photo_folders(photos_root)
    print(f"Photo folders mapped to refs: {len(ref_folders)}")

    SITE_LISTINGS_DIR.mkdir(parents=True, exist_ok=True)
    site_listings: list[dict] = []
    photo_report = []
    missing_photos = []

    for i, row in enumerate(unique, 1):
        ref = row["ref"]
        # Prefer excel photo folder if present
        folder = None
        excel_folder = row.get("photo_folder")
        if excel_folder:
            candidate = photos_root / str(excel_folder)
            if candidate.is_dir():
                folder = candidate
            else:
                # try normalized / mapped name
                mapped = ref_folders.get(norm_ref(excel_folder) or "")
                if mapped and mapped.is_dir():
                    folder = mapped
                else:
                    candidate = photos_root / str(norm_ref(excel_folder) or excel_folder)
                    if candidate.is_dir():
                        folder = candidate
        if folder is None:
            folder = ref_folders.get(ref)

        dest = SITE_LISTINGS_DIR / ref
        print(f"[{i}/{len(unique)}] ref {ref} ← {folder.name if folder else 'NO FOLDER'}")
        # Deduplicate only within a single listing so shared multi-ref folders
        # (e.g. 090–096) still attach photos to every unit.
        per_listing_hashes: set[str] = set()
        urls, meta = import_photos_for_ref(ref, folder, dest, args.max_photos, per_listing_hashes)
        photo_report.append(meta)
        if not urls:
            missing_photos.append(ref)
        site_listings.append(to_site_listing(row, urls))

    # Write site feed (array — required by listings.js)
    SITE_DATA.parent.mkdir(parents=True, exist_ok=True)
    SITE_DATA.write_text(json.dumps(site_listings, ensure_ascii=False, indent=2), encoding="utf-8")

    # Richer mirror next to source for reference
    mirror = source / "listings.site.json"
    mirror.write_text(
        json.dumps(
            {
                "generated": datetime.now(timezone.utc).isoformat(),
                "source_excel": str(xlsx),
                "count": len(site_listings),
                "listings": site_listings,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    report = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "source_excel": str(xlsx),
        "excel_rows": len(raw),
        "unique_listings": len(unique),
        "duplicate_rows": [
            {"ref": d["ref"], "kept": d["kept"], "dropped_price": d["dropped"].get("price")} for d in dups
        ],
        "listings_with_photos": sum(1 for L in site_listings if L["photo_count"] > 0),
        "listings_missing_photos": missing_photos,
        "total_images_imported": sum(L["photo_count"] for L in site_listings),
        "photo_report": photo_report,
        "type_counts": dict(Counter(L["property_type"] for L in site_listings)),
        "transaction_counts": dict(Counter(L["transaction"] for L in site_listings)),
        "output_json": str(SITE_DATA),
        "output_assets": str(SITE_LISTINGS_DIR),
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n=== IMPORT COMPLETE ===")
    print(f"Listings written: {len(site_listings)} → {SITE_DATA}")
    print(f"With photos: {report['listings_with_photos']} | Missing: {len(missing_photos)}")
    print(f"Images imported: {report['total_images_imported']}")
    print(f"Duplicates removed: {len(dups)}")
    if missing_photos:
        print("Missing photo refs:", ", ".join(missing_photos[:30]), ("..." if len(missing_photos) > 30 else ""))
    print(f"Report: {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
